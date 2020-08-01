from openpyxl import load_workbook
from openpyxl import Workbook
import googletrans
from pprint import pprint



wb = load_workbook('sample.xlsx')
sheet = wb.active 

nwb = Workbook() #開新excel
nws=nwb.active


translator = googletrans.Translator()


for row in sheet.rows:
    for cell in row:
        results = translator.translate(cell.value,dest='zh-tw').text
        cell.value=results

wb.save('sample.xlsx')
        





